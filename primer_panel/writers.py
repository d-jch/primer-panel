"""Output writers: BED, TSV, optional XLSX, FASTA, failed targets, primers.

Conceptual layers
-----------------
required_region  : raw CDS exons — must be covered by the amplicon.
extended_target  : required_region extended toward gene interior to product_min
                   (only when required_region < product_min).
design_template  : extended_target + primer_flank on each side — given to Primer3.
primer_product   : the actual PCR amplicon — must cover required_region.
"""

from __future__ import annotations

import csv
import logging
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING

from .config import PipelineConfig
from .cds_handler import CdsRequiredInterval
from .primer3_runner import PrimerResult

if TYPE_CHECKING:
    from .ensembl_client import TranscriptInfo


@dataclass
class Target:
    """A PCR target region covering one or more CDS required intervals.

    ``start`` / ``end`` define the **required_region** — the genomic interval
    that must be covered by any valid PCR product.  Downstream in build_records(),
    short required_regions are extended toward gene interior (→ extended_target)
    and then padded with primer_flank (→ design_template) for Primer3.
    """

    chrom: str
    start: int
    end: int
    strand: int
    needs_review: bool = False
    status: str = "ok"
    tiled: bool = False
    cds_exon_numbers: list[int] = field(default_factory=list)
    cds_exon_ids: list[str] = field(default_factory=list)
    cds_exon_coords: list[tuple[int, int]] = field(default_factory=list)

    @property
    def length(self) -> int:
        return self.end - self.start

logger = logging.getLogger(__name__)


@dataclass
class TargetRecord:
    """Flat record for one target, suitable for tabular output.

    Coordinates are split into four conceptual layers:
    required_region → extended_target → design_template.
    Stage 1 only produces these three layers; genomic product coordinates
    are computed by Stage 3 (in-silico PCR).
    """

    gene: str
    transcript_id: str
    selection_reason: str
    target_id: str
    # required_region (must be covered by PCR product)
    required_chrom: str
    required_start: int
    required_end: int
    required_length: int
    # extended_target (required_region extended toward gene interior to target-size)
    extended_chrom: str
    extended_start: int
    extended_end: int
    extended_length: int
    # design_template (extended_target + primer_flank; given to Primer3)
    template_chrom: str
    template_start: int
    template_end: int
    template_length: int
    # SEQUENCE_TARGET for Primer3 (extended_target relative to design_template)
    sequence_target_start_0based: int    # 0-based start within template
    sequence_target_length: int          # length of target within template
    sequence_target_for_primer3plus_1based: str  # "start,length" (1-based, for display)
    # metadata
    strand: str
    product_min: int
    product_max: int
    cds_exon_numbers: str    # comma-separated, e.g. "1,2,3"
    cds_exon_ids: str        # comma-separated
    covered_cds_count: int
    cds_exon_coords: str     # genomic coords "start1-end1,start2-end2"
    status: str
    needs_review: bool
    sequence_status: str     # "real" or "placeholder"
    target_qc_status: str    # Stage 1 QC: "ok" or description of issue


@dataclass
class FailedTarget:
    """Record for a gene that failed processing."""

    gene: str
    reason: str
    detail: str


# ------------------------------------------------------------------
# BED
# ------------------------------------------------------------------

def write_bed(targets: list[TargetRecord], path: Path) -> None:
    """Write targets in BED6 format using template coordinates."""
    with open(path, "w") as fh:
        for t in targets:
            fh.write(
                f"{t.template_chrom}\t{t.template_start}\t{t.template_end}"
                f"\t{t.target_id}\t0\t{t.strand}\n"
            )


def write_required_bed(targets: list[TargetRecord], path: Path) -> None:
    """Write required regions in BED6 format."""
    with open(path, "w") as fh:
        for t in targets:
            fh.write(
                f"{t.required_chrom}\t{t.required_start}\t{t.required_end}"
                f"\t{t.target_id}\t0\t{t.strand}\n"
            )


# ------------------------------------------------------------------
# TSV / XLSX summary
# ------------------------------------------------------------------

_SUMMARY_COLS = [
    "gene", "transcript_id", "selection_reason", "target_id",
    "required_chrom", "required_start", "required_end", "required_length",
    "extended_chrom", "extended_start", "extended_end", "extended_length",
    "template_chrom", "template_start", "template_end", "template_length",
    "sequence_target_start_0based", "sequence_target_length",
    "sequence_target_for_primer3plus_1based",
    "strand", "product_min", "product_max",
    "cds_exon_numbers", "cds_exon_ids", "covered_cds_count",
    "cds_exon_coords",
    "status", "needs_review", "sequence_status", "target_qc_status",
]


def write_summary_tsv(records: list[TargetRecord], path: Path) -> None:
    with open(path, "w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=_SUMMARY_COLS, delimiter="\t")
        writer.writeheader()
        for r in records:
            writer.writerow({col: getattr(r, col) for col in _SUMMARY_COLS})


def write_summary_xlsx(records: list[TargetRecord], path: Path) -> bool:
    """Write summary as Excel. Returns True on success, False if openpyxl missing."""
    try:
        from openpyxl import Workbook
    except ImportError:
        logger.info("openpyxl not installed — skipping XLSX output")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Targets"
    ws.append(_SUMMARY_COLS)
    for r in records:
        ws.append([getattr(r, col) for col in _SUMMARY_COLS])
    wb.save(path)
    return True


# ------------------------------------------------------------------
# Primers (Stage 2)
# ------------------------------------------------------------------

_PRIMER_COLS = [
    "target_id",
    "primer_rank", "forward_primer", "reverse_primer",
    "forward_tm", "reverse_tm", "tm_diff",
    "forward_gc", "reverse_gc",
    "primer_pair_penalty",
    "primer_left_start", "primer_left_len",
    "primer_right_start", "primer_right_len",
    "primer3_product_size",
    "primer3_status", "primer3_explain",
    "sequence_target_start_0based", "sequence_target_length",
]


@dataclass
class PrimerRecord:
    """Stage 2 output: Primer3 design result for one primer pair.

    Contains only Primer3 design-level fields.  Genomic product coordinates
    are produced by Stage 3 (in-silico PCR), not here.
    """

    target_id: str
    primer_rank: int
    forward_primer: str
    reverse_primer: str
    forward_tm: float
    reverse_tm: float
    tm_diff: float
    forward_gc: float
    reverse_gc: float
    primer_pair_penalty: float
    primer_left_start: int
    primer_left_len: int
    primer_right_start: int
    primer_right_len: int
    primer3_product_size: int       # Primer3-reported product size (template-relative)
    primer3_status: str
    primer3_explain: str
    sequence_target_start_0based: int
    sequence_target_length: int


def build_primer_records(
    records: list[TargetRecord],
    primer_results: dict[str, list[PrimerResult]],
    sequence_warnings: dict[str, str] | None = None,
) -> list[PrimerRecord]:
    """Merge target records with Primer3 results into flat PrimerRecord list.

    Stage 2 output: only Primer3 design-level fields.  No genomic product
    coordinates (those are produced by Stage 3).
    """
    if sequence_warnings is None:
        sequence_warnings = {}

    primer_records: list[PrimerRecord] = []

    for tr in records:
        results = primer_results.get(tr.target_id, [])

        if not results:
            primer_records.append(PrimerRecord(
                target_id=tr.target_id,
                primer_rank=0,
                forward_primer="",
                reverse_primer="",
                forward_tm=0.0,
                reverse_tm=0.0,
                tm_diff=0.0,
                forward_gc=0.0,
                reverse_gc=0.0,
                primer_pair_penalty=0.0,
                primer_left_start=0,
                primer_left_len=0,
                primer_right_start=0,
                primer_right_len=0,
                primer3_product_size=0,
                primer3_status="primer3_error",
                primer3_explain="no result from Primer3",
                sequence_target_start_0based=tr.sequence_target_start_0based,
                sequence_target_length=tr.sequence_target_length,
            ))
        else:
            for pr in results:
                if tr.strand == "-":
                    # Minus-strand gene: the design template is the + strand
                    # sequence, so Primer3's PRIMER_LEFT (binding the + strand)
                    # is the biological *reverse* primer, and PRIMER_RIGHT
                    # (binding the - strand) is the biological *forward*
                    # primer.  Swap labels so "forward" always means the
                    # transcription-direction primer.
                    # Coordinates stay as-is (both are + strand positions).
                    primer_records.append(PrimerRecord(
                        target_id=tr.target_id,
                        primer_rank=pr.primer_rank,
                        forward_primer=pr.reverse_primer,
                        reverse_primer=pr.forward_primer,
                        forward_tm=pr.reverse_tm,
                        reverse_tm=pr.forward_tm,
                        tm_diff=pr.tm_diff,
                        forward_gc=pr.reverse_gc,
                        reverse_gc=pr.forward_gc,
                        primer_pair_penalty=pr.primer_pair_penalty,
                        primer_left_start=pr.primer_left_start,
                        primer_left_len=pr.primer_left_len,
                        primer_right_start=pr.primer_right_start,
                        primer_right_len=pr.primer_right_len,
                        primer3_product_size=pr.primer3_product_size,
                        primer3_status=pr.primer3_status,
                        primer3_explain=pr.primer3_explain,
                        sequence_target_start_0based=pr.sequence_target_start_0based,
                        sequence_target_length=pr.sequence_target_length,
                    ))
                else:
                    primer_records.append(PrimerRecord(
                        target_id=tr.target_id,
                        primer_rank=pr.primer_rank,
                        forward_primer=pr.forward_primer,
                        reverse_primer=pr.reverse_primer,
                        forward_tm=pr.forward_tm,
                        reverse_tm=pr.reverse_tm,
                        tm_diff=pr.tm_diff,
                        forward_gc=pr.forward_gc,
                        reverse_gc=pr.reverse_gc,
                        primer_pair_penalty=pr.primer_pair_penalty,
                        primer_left_start=pr.primer_left_start,
                        primer_left_len=pr.primer_left_len,
                        primer_right_start=pr.primer_right_start,
                        primer_right_len=pr.primer_right_len,
                        primer3_product_size=pr.primer3_product_size,
                        primer3_status=pr.primer3_status,
                        primer3_explain=pr.primer3_explain,
                        sequence_target_start_0based=pr.sequence_target_start_0based,
                        sequence_target_length=pr.sequence_target_length,
                    ))

    return primer_records


def write_primers_tsv(records: list[PrimerRecord], path: Path) -> None:
    """Write primer results to TSV."""
    with open(path, "w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=_PRIMER_COLS, delimiter="\t")
        writer.writeheader()
        for r in records:
            writer.writerow({col: getattr(r, col) for col in _PRIMER_COLS})


def write_primers_xlsx(
    target_records: list[TargetRecord],
    primer_records: list[PrimerRecord],
    path: Path,
) -> bool:
    """Write combined Excel with Targets + Primers sheets.

    Returns True on success, False if openpyxl missing.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        logger.info("openpyxl not installed — skipping primers XLSX output")
        return False

    wb = Workbook()

    # Sheet 1: Targets
    ws_targets = wb.active
    ws_targets.title = "Targets"
    ws_targets.append(_SUMMARY_COLS)
    for r in target_records:
        ws_targets.append([getattr(r, col) for col in _SUMMARY_COLS])

    # Sheet 2: Primers
    ws_primers = wb.create_sheet("Primers")
    ws_primers.append(_PRIMER_COLS)
    for r in primer_records:
        ws_primers.append([getattr(r, col) for col in _PRIMER_COLS])

    wb.save(path)
    return True


# ------------------------------------------------------------------
# Failed targets
# ------------------------------------------------------------------

_FAILED_COLS = ["gene", "reason", "detail"]


def write_failed_targets(failed: list[FailedTarget], path: Path) -> None:
    """Write failed targets to TSV."""
    with open(path, "w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=_FAILED_COLS, delimiter="\t")
        writer.writeheader()
        for f in failed:
            writer.writerow({"gene": f.gene, "reason": f.reason, "detail": f.detail})


# ------------------------------------------------------------------
# FASTA
# ------------------------------------------------------------------

def write_fasta(
    records: list[TargetRecord],
    cfg: PipelineConfig,
    path: Path,
) -> str:
    """Write design_template FASTA.

    The sequence covers extended_target + primer_flank on each side.
    If cfg.genome_fasta is set and pyfaidx is available, extract real sequence.
    Otherwise write placeholder Ns.

    Returns sequence_status string: "real" or "placeholder".
    """
    if cfg.genome_fasta:
        return _write_fasta_real(records, cfg.genome_fasta, path)
    return _write_fasta_placeholder(records, path)


def _write_fasta_placeholder(records: list[TargetRecord], path: Path) -> str:
    with open(path, "w") as fh:
        for r in records:
            fh.write(
                f">{r.target_id} "
                f"template={r.template_chrom}:{r.template_start}-{r.template_end} "
                f"required={r.required_chrom}:{r.required_start}-{r.required_end} "
                f"strand={r.strand} placeholder\n"
            )
            # Write Ns in 80-char lines, length = template_length
            remaining = r.template_length
            while remaining > 0:
                chunk = min(80, remaining)
                fh.write("N" * chunk + "\n")
                remaining -= chunk
    return "placeholder"


def _write_fasta_real(
    records: list[TargetRecord],
    genome_fasta: Path,
    path: Path,
) -> str:
    try:
        from pyfaidx import Fasta
    except ImportError:
        logger.warning("pyfaidx not installed — falling back to placeholder sequences")
        return _write_fasta_placeholder(records, path)

    try:
        genome = Fasta(str(genome_fasta))
    except Exception as exc:
        logger.warning("Cannot open genome FASTA %s: %s — using placeholders", genome_fasta, exc)
        return _write_fasta_placeholder(records, path)

    with open(path, "w") as fh:
        for r in records:
            try:
                seq = str(genome[r.template_chrom][r.template_start:r.template_end])
            except Exception as exc:
                logger.warning(
                    "Cannot extract sequence for %s (%s:%d-%d): %s; using Ns",
                    r.target_id, r.template_chrom, r.template_start, r.template_end, exc,
                )
                seq = "N" * r.template_length
            fh.write(
                f">{r.target_id} "
                f"template={r.template_chrom}:{r.template_start}-{r.template_end} "
                f"required={r.required_chrom}:{r.required_start}-{r.required_end} "
                f"strand={r.strand}\n"
            )
            for i in range(0, len(seq), 80):
                fh.write(seq[i : i + 80] + "\n")
    return "real"


# ------------------------------------------------------------------
# High-level assembly
# ------------------------------------------------------------------

def _directional_extend(
    region_start: int,
    region_end: int,
    target_len: int,
    gene_required_start: int,
    gene_required_end: int,
) -> tuple[int, int]:
    """Extend *region* toward gene interior until its length reaches *target_len*.

    Returns (new_start, new_end).  The original region coordinates are NOT
    modified; the caller passes in the required_region bounds and gets back
    the extended bounds.

    Edge detection uses genome-coordinate proximity to the gene span
    (works for both + and − strand without strand-specific logic).

    If leftward extension is clamped at coordinate 0, the shortfall is
    compensated by extending further rightward so that the final length
    always equals *target_len* (unless the caller applies additional
    constraints).
    """
    cur_len = region_end - region_start
    if cur_len >= target_len:
        return region_start, region_end

    extra = target_len - cur_len

    dist_to_gene_start = region_start - gene_required_start
    dist_to_gene_end = gene_required_end - region_end
    gene_span = gene_required_end - gene_required_start
    edge_threshold = max(100, gene_span * 0.15)

    if dist_to_gene_start <= edge_threshold and dist_to_gene_end > edge_threshold:
        # At gene start edge → extend rightward (toward gene interior)
        return region_start, region_end + extra
    elif dist_to_gene_end <= edge_threshold and dist_to_gene_start > edge_threshold:
        # At gene end edge → extend leftward (toward gene interior)
        new_start = max(0, region_start - extra)
        # If leftward was clamped at 0, compensate rightward
        left_actual = region_start - new_start
        shortfall = extra - left_actual
        return new_start, region_end + shortfall
    else:
        # Middle or both edges → symmetric
        left = extra // 2
        right = extra - left
        new_start = max(0, region_start - left)
        left_actual = region_start - new_start
        shortfall = left - left_actual
        return new_start, region_end + right + shortfall


def build_records(
    gene: str,
    ti: TranscriptInfo,
    targets: list[Target],
    cfg: PipelineConfig,
    sequence_status: str,
    gene_required_start: int | None = None,
    gene_required_end: int | None = None,
) -> list[TargetRecord]:
    """Convert internal targets to flat TargetRecord list.

    For each target (whose start/end = required_region from raw CDS):

    1. required_region = original CDS coordinates (unchanged).
    2. If required_length < product_min, extend toward gene interior to
       reach product_min → extended_target.
    3. Add primer_flank on both sides of the extended target → design_template.
    4. Clamp start to >= 0.
    5. Compute SEQUENCE_TARGET relative coordinates for Primer3.

    The ``required_*`` fields in the output reflect the original CDS
    coordinates.  ``template_*`` fields reflect the final design_template
    given to Primer3.
    """
    if gene_required_start is None:
        gene_required_start = min(t.start for t in targets) if targets else 0
    if gene_required_end is None:
        gene_required_end = max(t.end for t in targets) if targets else 0

    records: list[TargetRecord] = []
    for t in targets:
        if t.cds_exon_numbers:
            if len(t.cds_exon_numbers) == 1:
                target_id = f"{gene}_cds{t.cds_exon_numbers[0]}"
            else:
                target_id = f"{gene}_cds{t.cds_exon_numbers[0]}_{t.cds_exon_numbers[-1]}"
        else:
            target_id = f"{gene}_t{len(records) + 1}"

        coords_str = ",".join(f"{s}-{e}" for s, e in t.cds_exon_coords)

        # --- Step 1: extend required_region toward gene interior to product_min ---
        ext_start, ext_end = _directional_extend(
            t.start, t.end, cfg.product_min,
            gene_required_start, gene_required_end,
        )

        # --- Step 2: add primer_flank on both sides → design_template ---
        desired_template_len = (ext_end - ext_start) + 2 * cfg.primer_flank
        template_start = max(0, ext_start - cfg.primer_flank)
        template_end = ext_end + cfg.primer_flank
        # If clamping start to 0 caused shortfall, compensate on the right
        actual_len = template_end - template_start
        if actual_len < desired_template_len and template_start == 0:
            template_end += (desired_template_len - actual_len)
        template_length = template_end - template_start

        # --- Step 3: compute SEQUENCE_TARGET for Primer3 ---
        # extended_target relative to design_template (0-based)
        seq_target_start_0based = ext_start - template_start
        seq_target_length = ext_end - ext_start
        seq_target_1based = f"{seq_target_start_0based + 1},{seq_target_length}"

        # --- Step 4: Stage 1 QC ---
        qc_issues: list[str] = []
        if seq_target_start_0based < 0:
            qc_issues.append("target_starts_before_template")
        if seq_target_start_0based + seq_target_length > template_length:
            qc_issues.append("target_ends_after_template")
        if ext_start > t.start:
            qc_issues.append("extended_does_not_cover_required_start")
        if ext_end < t.end:
            qc_issues.append("extended_does_not_cover_required_end")
        for cds_start, cds_end in t.cds_exon_coords:
            if cds_start < ext_start or cds_end > ext_end:
                qc_issues.append(f"cds_exon_{cds_start}_{cds_end}_not_covered_by_extended")
                break
        if template_length < ext_end - ext_start + 2 * cfg.primer_flank:
            qc_issues.append("template_too_short_for_extended_plus_flank")
        if t.tiled:
            qc_issues.append("tiled")
        if sequence_status == "placeholder":
            qc_issues.append("placeholder_sequence")
        if sequence_status == "real":
            # Sequence-level checks are done after FASTA extraction;
            # for now just record the base status
            pass

        target_qc = ";".join(qc_issues) if qc_issues else "ok"

        # --- Status ---
        needs_review = t.needs_review
        status = t.status
        if template_length < cfg.product_min:
            needs_review = True
            status = "template_short"
        elif status == "ok":
            status = "template_supports_product_max" if template_length >= cfg.product_max else "template_ok"

        records.append(TargetRecord(
            gene=gene,
            transcript_id=ti.transcript_id,
            selection_reason=ti.selection_reason,
            target_id=target_id,
            required_chrom=t.chrom,
            required_start=t.start,
            required_end=t.end,
            required_length=t.length,
            extended_chrom=t.chrom,
            extended_start=ext_start,
            extended_end=ext_end,
            extended_length=ext_end - ext_start,
            template_chrom=t.chrom,
            template_start=template_start,
            template_end=template_end,
            template_length=template_length,
            sequence_target_start_0based=seq_target_start_0based,
            sequence_target_length=seq_target_length,
            sequence_target_for_primer3plus_1based=seq_target_1based,
            strand="+" if t.strand == 1 else "-",
            product_min=cfg.product_min,
            product_max=cfg.product_max,
            cds_exon_numbers=",".join(str(n) for n in t.cds_exon_numbers),
            cds_exon_ids=",".join(t.cds_exon_ids),
            covered_cds_count=len(t.cds_exon_numbers),
            cds_exon_coords=coords_str,
            status=status,
            needs_review=needs_review,
            sequence_status=sequence_status,
            target_qc_status=target_qc,
        ))
    return records


# ------------------------------------------------------------------
# QC summary
# ------------------------------------------------------------------

def _detect_sequence_status(cfg: PipelineConfig) -> str:
    """Detect actual sequence status by checking the FASTA file content.

    Returns 'real' if targets.fa contains at least one non-placeholder sequence,
    'placeholder' if all sequences are N, or 'unknown' if file doesn't exist.
    """
    fa_path = cfg.output_dir / "targets.fa"
    if not fa_path.exists():
        return "unknown"
    try:
        has_real = False
        has_placeholder = False
        with open(fa_path) as fh:
            for line in fh:
                if line.startswith(">"):
                    continue
                seq = line.strip()
                if not seq:
                    continue
                if all(c in ("N", "n") for c in seq):
                    has_placeholder = True
                else:
                    has_real = True
        if has_real:
            return "real"
        if has_placeholder:
            return "placeholder"
        return "unknown"
    except Exception:
        return "unknown"


def write_qc_summary(
    records: list[TargetRecord],
    failed: list[FailedTarget],
    primer_records: list[PrimerRecord] | None,
    cfg: PipelineConfig,
    path: Path,
) -> None:
    """Write a run summary with QC metrics."""
    lines: list[str] = []

    # Detect actual sequence status from FASTA file (more reliable than records)
    actual_seq_status = _detect_sequence_status(cfg)

    # --- Stage 1 stats ---
    genes = sorted(set(r.gene for r in records))
    lines.append(f"gene_count\t{len(genes)}")
    lines.append(f"target_count\t{len(records)}")

    targets_by_gene = Counter(r.gene for r in records)
    lines.append(f"targets_by_gene\t{', '.join(f'{g}:{targets_by_gene[g]}' for g in genes)}")

    if records:
        req_lens = [r.required_length for r in records]
        ext_lens = [r.extended_length for r in records]
        tpl_lens = [r.template_length for r in records]

        lines.append(f"required_length_min\t{min(req_lens)}")
        lines.append(f"required_length_max\t{max(req_lens)}")
        lines.append(f"extended_length_min\t{min(ext_lens)}")
        lines.append(f"extended_length_max\t{max(ext_lens)}")
        lines.append(f"template_length_min\t{min(tpl_lens)}")
        lines.append(f"template_length_max\t{max(tpl_lens)}")

        tpl_below_max = sum(1 for t in tpl_lens if t < cfg.product_max)
        lines.append(f"template_length_below_product_max\t{tpl_below_max}")

        needs_review_count = sum(1 for r in records if r.needs_review)
        lines.append(f"needs_review_count\t{needs_review_count}")

        # Use detected sequence status (from FASTA content) as primary source
        if actual_seq_status != "unknown":
            lines.append(f"sequence_status_{actual_seq_status}\t{len(records)}")
        else:
            # Fallback to records
            seq_statuses = Counter(r.sequence_status for r in records)
            for ss, cnt in seq_statuses.items():
                lines.append(f"sequence_status_{ss}\t{cnt}")

    lines.append(f"failed_gene_count\t{len(failed)}")
    for f in failed:
        lines.append(f"failed_gene\t{f.gene}\t{f.reason}\t{f.detail}")

    # --- Stage 2 stats ---
    if primer_records is not None:
        all_tids = sorted(set(pr.target_id for pr in primer_records))
        lines.append(f"primer_target_count\t{len(all_tids)}")

        ok_pairs = sum(1 for pr in primer_records if pr.primer3_status == "ok")
        lines.append(f"primer_ok_pair_count\t{ok_pairs}")

        records_by_target: dict[str, list[PrimerRecord]] = defaultdict(list)
        for pr in primer_records:
            records_by_target[pr.target_id].append(pr)

        # Per-target analysis
        per_target_ok: dict[str, int] = {}
        per_target_best_size: dict[str, int] = {}
        per_target_best_penalty: dict[str, float] = {}
        per_target_reason: dict[str, str] = {}

        for tid in all_tids:
            tid_records = records_by_target[tid]
            ok_recs = [pr for pr in tid_records if pr.primer3_status == "ok"]
            per_target_ok[tid] = len(ok_recs)
            if ok_recs:
                best = min(ok_recs, key=lambda r: r.primer_pair_penalty)
                per_target_best_size[tid] = best.primer3_product_size
                per_target_best_penalty[tid] = best.primer_pair_penalty
                per_target_reason[tid] = ""
            else:
                per_target_best_size[tid] = 0
                per_target_best_penalty[tid] = 0.0
                # Collect unique failure reasons
                reasons = set()
                for pr in tid_records:
                    if pr.primer3_explain:
                        reasons.add(pr.primer3_explain)
                per_target_reason[tid] = "; ".join(sorted(reasons)) if reasons else pr.primer3_status

        ok_targets = sum(1 for cnt in per_target_ok.values() if cnt > 0)
        failed_targets = len(all_tids) - ok_targets
        lines.append(f"primer_ok_target_count\t{ok_targets}")
        lines.append(f"primer_failed_target_count\t{failed_targets}")

        status_counts = Counter(pr.primer3_status for pr in primer_records)
        for st, cnt in sorted(status_counts.items()):
            lines.append(f"primer3_status_{st}\t{cnt}")

        # Per-target details
        lines.append("")
        lines.append("# Per-target primer details")
        lines.append("target_id\tok_count\tbest_product_size\tbest_penalty\tfailure_reason")
        for tid in all_tids:
            lines.append(
                f"{tid}\t{per_target_ok[tid]}\t{per_target_best_size[tid]}"
                f"\t{per_target_best_penalty[tid]}\t{per_target_reason[tid]}"
            )

    with open(path, "w") as fh:
        fh.write("\n".join(lines) + "\n")


# ------------------------------------------------------------------
# Stage 3: Specificity output
# ------------------------------------------------------------------

_SPECIFICITY_COLS = [
    "target_id", "primer_rank",
    "forward_primer", "reverse_primer",
    "forward_tm", "reverse_tm", "tm_diff",
    "forward_gc", "reverse_gc",
    "primer_pair_penalty",
    "primer3_product_size",
    "primer3_status", "primer3_explain",
    "sequence_target_start_0based", "sequence_target_length",
    # Stage 3 specificity fields
    "insilico_status", "insilico_hit_count", "insilico_hits",
    "insilico_best_chrom", "insilico_best_start", "insilico_best_end",
    "insilico_best_size", "specificity_pass",
    "expected_target_chrom", "expected_target_start", "expected_target_end",
    "specificity_explain",
    # Common dbSNP annotation fields
    "common_snp_risk", "left_primer_common_snp_count",
    "right_primer_common_snp_count", "left_primer_3p_common_snp_count",
    "right_primer_3p_common_snp_count", "common_snp_hits",
]


@dataclass
class SpecificityRecord:
    """Stage 3 output: primer pair with in-silico PCR specificity info.

    Combines Primer2 design fields with Stage 3 genomic product coordinates.
    """

    # Stage 2 PrimerRecord fields
    target_id: str
    primer_rank: int
    forward_primer: str
    reverse_primer: str
    forward_tm: float
    reverse_tm: float
    tm_diff: float
    forward_gc: float
    reverse_gc: float
    primer_pair_penalty: float
    primer_left_start: int
    primer_left_len: int
    primer_right_start: int
    primer_right_len: int
    primer3_product_size: int
    primer3_status: str
    primer3_explain: str
    sequence_target_start_0based: int
    sequence_target_length: int
    # Stage 3 specificity fields
    insilico_status: str
    insilico_hit_count: int
    insilico_hits: str
    insilico_best_chrom: str
    insilico_best_start: int
    insilico_best_end: int
    insilico_best_size: int
    specificity_pass: bool
    expected_target_chrom: str
    expected_target_start: int
    expected_target_end: int
    specificity_explain: str
    # Common dbSNP annotation fields
    common_snp_risk: str = "none"
    left_primer_common_snp_count: int = 0
    right_primer_common_snp_count: int = 0
    left_primer_3p_common_snp_count: int = 0
    right_primer_3p_common_snp_count: int = 0
    common_snp_hits: str = ""


def build_specificity_records(
    primer_records: list[PrimerRecord],
    specificity_results: dict[str, object],  # primer_name -> SpecificityResult
    expected_coords: dict[str, tuple[str, int, int]] | None = None,  # target_id -> (chrom, start, end)
    snp_annotations: dict[str, dict] | None = None,  # primer_name -> SNP annotation dict
) -> list[SpecificityRecord]:
    """Merge primer records with Stage 3 specificity results.

    Args:
        primer_records: Stage 2 output
        specificity_results: {primer_name: SpecificityResult}
        expected_coords: {target_id: (chrom, start, end)} for expected target region
        snp_annotations: {primer_name: annotation_dict} for common dbSNP
    """
    if expected_coords is None:
        expected_coords = {}
    if snp_annotations is None:
        snp_annotations = {}

    records: list[SpecificityRecord] = []
    for pr in primer_records:
        primer_name = f"{pr.target_id}_rank{pr.primer_rank}"
        sp = specificity_results.get(primer_name)

        exp_chrom, exp_start, exp_end = expected_coords.get(
            pr.target_id, ("", 0, 0)
        )

        snp = snp_annotations.get(primer_name, {})
        snp_fields = {
            "common_snp_risk": snp.get("common_snp_risk", "none"),
            "left_primer_common_snp_count": snp.get("left_primer_common_snp_count", 0),
            "right_primer_common_snp_count": snp.get("right_primer_common_snp_count", 0),
            "left_primer_3p_common_snp_count": snp.get("left_primer_3p_common_snp_count", 0),
            "right_primer_3p_common_snp_count": snp.get("right_primer_3p_common_snp_count", 0),
            "common_snp_hits": snp.get("common_snp_hits", ""),
        }

        primer_fields = {k: getattr(pr, k) for k in PrimerRecord.__dataclass_fields__}
        common = {
            "expected_target_chrom": exp_chrom,
            "expected_target_start": exp_start,
            "expected_target_end": exp_end,
            **snp_fields,
        }

        if sp is None:
            records.append(SpecificityRecord(
                **primer_fields,
                insilico_status="not_checked",
                insilico_hit_count=0,
                insilico_hits="",
                insilico_best_chrom="",
                insilico_best_start=0,
                insilico_best_end=0,
                insilico_best_size=0,
                specificity_pass=False,
                specificity_explain="primer3_status != ok; not checked",
                **common,
            ))
        else:
            records.append(SpecificityRecord(
                **primer_fields,
                insilico_status=sp.insilico_status,
                insilico_hit_count=sp.insilico_hit_count,
                insilico_hits=sp.insilico_hits,
                insilico_best_chrom=sp.insilico_best_chrom,
                insilico_best_start=sp.insilico_best_start,
                insilico_best_end=sp.insilico_best_end,
                insilico_best_size=sp.insilico_best_size,
                specificity_pass=sp.specificity_pass,
                specificity_explain=sp.specificity_explain,
                **common,
            ))
    return records


def write_specificity_tsv(records: list[SpecificityRecord], path: Path) -> None:
    """Write specificity results to TSV."""
    with open(path, "w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=_SPECIFICITY_COLS, delimiter="\t")
        writer.writeheader()
        for r in records:
            writer.writerow({col: getattr(r, col) for col in _SPECIFICITY_COLS})


def write_unique_primers(records: list[SpecificityRecord], path: Path) -> None:
    """Write only unique_pass primers to TSV."""
    with open(path, "w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=_SPECIFICITY_COLS, delimiter="\t")
        writer.writeheader()
        for r in records:
            if r.primer3_status == "ok" and r.specificity_pass:
                writer.writerow({col: getattr(r, col) for col in _SPECIFICITY_COLS})


_SPECIFICITY_CLEAN_COLS = [
    "target_id", "primer_rank",
    "forward_primer", "reverse_primer",
    "forward_tm", "reverse_tm",
    "forward_gc", "reverse_gc",
    "primer3_product_size",
    "insilico_hits",
]


def write_specificity_clean_tsv(records: list[SpecificityRecord], path: Path) -> None:
    """Write a clean specificity TSV: unique_pass only, no SNP overlap, minimal columns.

    Filters to rows where insilico_status == "unique_pass" and
    common_snp_risk == "none", then drops all columns except the 10
    retained in _SPECIFICITY_CLEAN_COLS.
    """
    with open(path, "w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=_SPECIFICITY_CLEAN_COLS, delimiter="\t")
        writer.writeheader()
        for r in records:
            if r.insilico_status == "unique_pass" and r.common_snp_risk == "none":
                writer.writerow({col: getattr(r, col) for col in _SPECIFICITY_CLEAN_COLS})


def write_specificity_summary(
    records: list[SpecificityRecord],
    path: Path,
) -> None:
    """Write Stage 3 specificity summary."""
    lines: list[str] = []

    # Only consider records that were checked (primer3_status == ok)
    checked = [r for r in records if r.primer3_status == "ok"]
    unique_pass = [r for r in checked if r.specificity_pass]
    multi_hit = [r for r in checked if r.insilico_status == "multi_hit"]
    no_hit = [r for r in checked if r.insilico_status == "no_hit"]
    off_target = [r for r in checked if r.insilico_status == "unique_off_target"]
    pcr_error = [r for r in checked if r.insilico_status == "pcr_error"]

    lines.append(f"total_ok_primers_input\t{len(checked)}")
    lines.append(f"unique_pass_count\t{len(unique_pass)}")
    lines.append(f"multi_hit_count\t{len(multi_hit)}")
    lines.append(f"no_hit_count\t{len(no_hit)}")
    lines.append(f"unique_off_target_count\t{len(off_target)}")
    lines.append(f"pcr_error_count\t{len(pcr_error)}")

    # Per-target
    all_tids = sorted(set(r.target_id for r in checked))
    targets_with_unique = sum(
        1 for tid in all_tids
        if any(r.specificity_pass and r.target_id == tid for r in checked)
    )
    lines.append(f"targets_with_unique_primer\t{targets_with_unique}")
    lines.append(f"targets_without_unique_primer\t{len(all_tids) - targets_with_unique}")

    lines.append("")
    lines.append("# Per-target specificity details")
    lines.append("target_id\tunique_count\tbest_unique_penalty\tbest_unique_product_size")
    for tid in all_tids:
        tid_unique = [r for r in unique_pass if r.target_id == tid]
        if tid_unique:
            best = min(tid_unique, key=lambda r: r.primer_pair_penalty)
            lines.append(f"{tid}\t{len(tid_unique)}\t{best.primer_pair_penalty}\t{best.primer3_product_size}")
        else:
            lines.append(f"{tid}\t0\t\t")

    with open(path, "w") as fh:
        fh.write("\n".join(lines) + "\n")
